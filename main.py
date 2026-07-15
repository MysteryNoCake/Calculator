import openpyxl
import numpy as np
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import os
import copy
import warnings

warnings.filterwarnings('ignore', message='Data Validation extension is not supported')
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')


class PayrollCalculator:
    """Калькулятор расчета фонда оплаты труда (ФОТ)"""

    def __init__(self):
        self.employees_data = []
        self.original_employees_data = []
        self.modified_employees = {}
        self.reference_data = {}
        self.motivation_map = {}
        self.changes_log = []
        self.base_fot_total = 0.0  # ✅ ЗАФИКСИРОВАННЫЙ ФОТ ИЗ СТОЛБЦА FN
        self.additional_positions = []  # ✅ Список дополнительных позиций
        self.hypotheses_data = {
            'working_hours': {}, 'staffing': {}, 'tariff_coefficients': {},
            'field_coefficients': {}, 'additional_coefficients': {}, 'bonus_payment': {},
            'night_share': 0.267,
            'sv_rates': {}
        }
        self.months = ['январь', 'февраль', 'март', 'апрель', 'май', 'июнь',
                       'июль', 'август', 'сентябрь', 'октябрь', 'ноябрь', 'декабрь']
        self.months_short = ['янв', 'фев', 'мар', 'апр', 'май', 'июн',
                             'июл', 'авг', 'сен', 'окт', 'ноя', 'дек']
        self.year = 2026
        self.sv_thresholds = {
            'threshold1': 2225000, 'threshold2': 2959000,
            'rate1': 0.302, 'rate2': 0.151, 'rate3': 0.051
        }
        self.total_res_a_excel = 0.0

    def _parse_number(self, value):
        """Точный парсинг чисел из Excel"""
        if value is None:
            return 0.0
        if isinstance(value, (int, float)):
            return 0.0 if np.isnan(value) or np.isinf(value) else float(value)
        s = str(value).strip()
        if not s:
            return 0.0
        is_percent = s.endswith('%')
        if is_percent:
            s = s[:-1].strip()
        for ch in ('\u00a0', '\u202f', '\u2009', '\u2007', ' '):
            s = s.replace(ch, '')
        s = s.replace('#REF!', '0').replace('ERROR', '0')
        if not s or s in ('-', '—'):
            return 0.0
        try:
            if ',' in s and '.' not in s:
                left, _, right = s.rpartition(',')
                if right.isdigit() and len(right) <= 2 and left.replace('.', '').replace('-', '').isdigit():
                    s = left + '.' + right
                else:
                    s = s.replace(',', '')
            elif ',' in s and '.' in s:
                s = s.replace('.', '').replace(',', '.') if s.rfind(',') > s.rfind('.') else s.replace(',', '')
            return float(s) / 100.0 if is_percent else float(s)
        except (TypeError, ValueError):
            return 0.0

    def load_employee_data(self, filepath):
        """Загрузка данных из Excel"""
        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
            if 'Справочники' in wb.sheetnames:
                self.reference_data = self._parse_directory_sheet(wb['Справочники'])
            employee_sheet = next((s for s in wb.sheetnames if s.upper() == 'ФОРМАТ'), wb.sheetnames[0])
            ws = wb[employee_sheet]
            self._parse_employee_sheet(ws)
            self.original_employees_data = [copy.deepcopy(e) for e in self.employees_data]
            self.modified_employees = {}
            self.changes_log = []
            self.additional_positions = []  # ✅ Сброс при загрузке
            ws_hyp = next((s for s in wb.sheetnames if 'Гипотез' in s), None)
            if ws_hyp:
                self.hypotheses_data.update(self._parse_hypotheses_sheet(wb[ws_hyp]))
            return self
        except Exception as e:
            raise Exception(f"Ошибка загрузки файла: {str(e)}")

    def _parse_directory_sheet(self, ws):
        """Парсинг справочников"""
        directories = {'legal_entities': {}, 'cities': {}}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not any(row):
                continue
            d = dict(zip([str(c) if c else '' for c in ws[1]], row))
            motiv_key = str(d.get('Мес. Мотивация', '')).strip()
            motiv_val = self._parse_number(d.get('Премия vs Оклад', 0.0))
            if motiv_key:
                self.motivation_map[motiv_key] = motiv_val
            if d.get('Юрлицо'):
                directories['legal_entities'][d['Юрлицо']] = {'sv_group': d.get('Группа для СВ', 'А')}
            if d.get('Город базирования'):
                directories['cities'][d['Город базирования']] = {'schedule': d.get('График работы', '')}
        return directories

    def _parse_employee_sheet(self, ws):
        """Парсинг сотрудников"""
        employees = []
        current_dir = ''
        self.total_res_a_excel = 0.0
        self.base_fot_total = 0.0  # ✅ СБРОС ПЕРЕД ЗАГРУЗКОЙ
        for row in ws.iter_rows(min_row=6, values_only=True):
            if not row or not any(row):
                continue
            col_a = row[0] if len(row) > 0 else None
            if col_a and 'ДИРЕКЦИЯ' in str(col_a).upper():
                current_dir = str(col_a).strip()
                if not (len(row) > 8 and row[8] and str(row[8]).strip()):
                    continue
            fio = row[8] if len(row) > 8 else None
            position = row[7] if len(row) > 7 else (row[8] if len(row) > 8 else None)
            if not fio or str(fio).strip() == '':
                continue
            fio_str = str(fio).strip().upper()
            if any(x in fio_str for x in ['ФИО', 'ДОЛЖНОСТЬ', 'ДИРЕКЦИЯ', 'СТОЛБЕЦ', 'ИТОГО']):
                continue
            base_salary_monthly = [self._parse_number(row[i]) if len(row) > i else 0.0 for i in range(61, 73)]
            if all(v == 0.0 for v in base_salary_monthly):
                base_sal = self._parse_number(row[14]) if len(row) > 14 else 0.0
                if base_sal == 0.0 and len(row) > 15:
                    base_sal = self._parse_number(row[15])
                base_salary_monthly = [base_sal] * 12
            motiv_raw = str(row[39]).strip() if len(row) > 39 else ''
            monthly_bonus_pct = self.motivation_map.get(motiv_raw,
                                                        self._parse_number(row[39]) if len(row) > 39 else 0.0)
            headcounts = [self._parse_number(row[i]) if len(row) > i else 1.0 for i in range(47, 59)]
            headcounts = [v if v > 0 else 1.0 for v in headcounts]
            compensation_summer = self._parse_number(row[37]) if len(row) > 37 else 0.0
            compensation_winter = self._parse_number(row[38]) if len(row) > 38 else 0.0
            sv_excel = [round(self._parse_number(row[i]), 2) for i in range(103, 115)]
            sv_total_excel = self._parse_number(row[115]) if len(row) > 115 else 0.0
            comp_excel = [self._parse_number(row[i]) if len(row) > i else 0.0 for i in range(116, 129)]
            comp_total_excel = self._parse_number(row[129]) if len(row) > 129 else 0.0
            res_q_excel = [self._parse_number(row[i]) if len(row) > i else 0.0 for i in range(130, 142)]
            res_q_total_excel = self._parse_number(row[142]) if len(row) > 142 else 0.0
            res_a_excel = [self._parse_number(row[i]) if len(row) > i else 0.0 for i in range(143, 155)]
            res_a_total_excel = self._parse_number(row[155]) if len(row) > 155 else 0.0
            self.total_res_a_excel += res_a_total_excel
            # ✅ ЧТЕНИЕ БАЗОВОГО ФОТ ИЗ СТОЛБЦА FN (индекс 169)
            base_fot = self._parse_number(row[169]) if len(row) > 169 else 0.0
            self.base_fot_total += base_fot  # ✅ НАКАПЛИВАЕМ СУММУ
            legal_entity = 'АСГ'
            for row_dir in ws.iter_rows(min_row=2, max_row=5, values_only=True):
                if row_dir and len(row_dir) > 1:
                    dir_name = str(row_dir[0]) if row_dir[0] else ''
                    if 'ДИРЕКЦИЯ' in dir_name.upper():
                        if len(row_dir) > 1 and row_dir[1]:
                            legal_entity = str(row_dir[1]).strip()
                        break
            emp = {
                'directorate': current_dir, 'fio': str(fio).strip(),
                'position': str(position).strip() if position else '',
                'base_salary': base_salary_monthly[0] if base_salary_monthly else 0.0,
                'base_salary_monthly': list(base_salary_monthly),
                'allowance_cat1': self._parse_number(row[16]) if len(row) > 16 else 0.0,
                'allowance_cat2': self._parse_number(row[17]) if len(row) > 17 else 0.0,
                'travel_allowance': self._parse_number(row[18]) if len(row) > 18 else 0.0,
                'rk_percent': self._parse_number(row[19]) if len(row) > 19 else 0.0,
                'sn_percent': self._parse_number(row[20]) if len(row) > 20 else 0.0,
                'harmfulness_percent': self._parse_number(row[21]) if len(row) > 21 else 0.0,
                'night_percent': self._parse_number(row[22]) if len(row) > 22 else 0.0,
                'equipment_compensation': self._parse_number(row[25]) if len(row) > 25 else 0.0,
                'car_rent': self._parse_number(row[26]) if len(row) > 26 else 0.0,
                'crew_rent': self._parse_number(row[27]) if len(row) > 27 else 0.0,
                'fuel_limit_summer': self._parse_number(row[28]) if len(row) > 28 else 0.0,
                'fuel_limit_winter': self._parse_number(row[29]) if len(row) > 29 else 0.0,
                'transport_expenses': self._parse_number(row[30]) if len(row) > 30 else 0.0,
                'car_compensation': self._parse_number(row[31]) if len(row) > 31 else 0.0,
                'phone_limit': self._parse_number(row[32]) if len(row) > 32 else 0.0,
                'internet_limit': self._parse_number(row[33]) if len(row) > 33 else 0.0,
                'dms': self._parse_number(row[34]) if len(row) > 34 else 0.0,
                'housing_compensation': self._parse_number(row[35]) if len(row) > 35 else 0.0,
                'travel_compensation': self._parse_number(row[36]) if len(row) > 36 else 0.0,
                'compensation_summer': compensation_summer,
                'compensation_winter': compensation_winter,
                'monthly_bonus_percent': monthly_bonus_pct,
                'quarterly_bonus_percent': self._parse_number(row[40]) if len(row) > 40 else 0.0,
                'annual_bonus_percent': self._parse_number(row[41]) if len(row) > 41 else 0.0,
                'salary_type': str(row[13]).lower().strip() if len(row) > 13 else '',
                'last_employee': str(row[10]).strip().lower() if len(row) > 10 else '',
                'headcounts': list(headcounts),
                'legal_entity': legal_entity,
                'excel_sv_monthly': list(sv_excel), 'excel_sv_total': sv_total_excel,
                'excel_comp_monthly': list(comp_excel), 'excel_comp_total': comp_total_excel,
                'excel_res_q_monthly': list(res_q_excel), 'excel_res_q_total': res_q_total_excel,
                'excel_res_a_monthly': list(res_a_excel), 'excel_res_a_total': res_a_total_excel,
                'base_fot_year': base_fot,
                'gross_additions': 0.0,
                'comment': ''
            }
            employees.append(emp)
        self.employees_data = employees

    def _parse_hypotheses_sheet(self, ws):
        """Парсинг гипотез"""
        hyp = {'tariff_coefficients': {}, 'field_coefficients': {}, 'additional_coefficients': {},
               'bonus_payment': {}, 'staffing': {}, 'sv_rates': {}}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not any(row):
                continue
            d = dict(zip([str(c) if c else '' for c in ws[1]], row))
            if d.get('Год') == '2026' and d.get('Месяц'):
                m = int(self._parse_number(d.get('Номер месяца', 0)))
                if 1 <= m <= 12:
                    hyp['staffing'][m] = self._parse_number(d.get('Укомплект-ть ЭТК', 0.9))
                    hyp['tariff_coefficients'][m] = self._parse_number(d.get('Коэффициент тарифа', 1.0))
                    hyp['field_coefficients'][m] = self._parse_number(d.get('Полевой коэффициент', 1.0))
                    hyp['additional_coefficients'][m] = self._parse_number(d.get('Дополнительный коэффициент', 1.0))
            if 'ночных' in str(d.get('Месяц2', '')).lower():
                hyp['night_share'] = self._parse_number(d.get('Доля ночных', 0.267))
            q = str(d.get('Месяц', ''))
            rate = self._parse_number(d.get('Процент выплаты кв. премии')) if d.get(
                'Процент выплаты кв. премии') else None
            if rate:
                if '1 кв' in q:
                    hyp['bonus_payment']['q1'] = rate
                elif '2 кв' in q:
                    hyp['bonus_payment']['q2'] = rate
                elif '3 кв' in q:
                    hyp['bonus_payment']['q3'] = rate
                elif '4 кв' in q:
                    hyp['bonus_payment']['q4'] = rate
            admin_rate = self._parse_number(d.get('Админ.', 0.0))
            oper_rate = self._parse_number(d.get('Операции', 0.0))
            col_a = self._parse_number(d.get('A', 0.0))
            col_o = self._parse_number(d.get('O', 0.0))
            if admin_rate > 0:
                hyp['sv_rates'][m] = {'admin': admin_rate, 'oper': oper_rate, 'A': col_a, 'O': col_o}
        self.hypotheses_data.update(hyp)
        return hyp

    def calculate_fot_for_position(self, position_data):
        """Расчет ФОТ для новой позиции"""
        base_salary = position_data.get('base_salary', 0.0)
        rk_percent = position_data.get('rk_percent', 0.0)
        sn_percent = position_data.get('sn_percent', 0.0)
        monthly_bonus_pct = position_data.get('monthly_bonus_percent', 0.0)
        quarterly_bonus_pct = position_data.get('quarterly_bonus_percent', 0.0)
        annual_bonus_pct = position_data.get('annual_bonus_percent', 0.0)
        compensation_summer = position_data.get('compensation_summer', 0.0)
        compensation_winter = position_data.get('compensation_winter', 0.0)
        months_count = position_data.get('months_count', 12)
        monthly_salary = base_salary * (1 + rk_percent / 100) * (1 + sn_percent / 100)
        monthly_bonus = monthly_salary * (monthly_bonus_pct / 100)
        monthly_gross = monthly_salary + monthly_bonus
        quarterly_reserve = monthly_salary * (quarterly_bonus_pct / 100) / 3
        annual_reserve = monthly_salary * (annual_bonus_pct / 100) / 12
        sv_monthly = monthly_gross * 0.302
        compensation = (compensation_summer + compensation_winter) / 2
        monthly_fot = monthly_gross + sv_monthly + quarterly_reserve + annual_reserve + compensation
        total_fot = monthly_fot * months_count
        return {
            'monthly_salary': monthly_salary,
            'monthly_bonus': monthly_bonus,
            'monthly_gross': monthly_gross,
            'quarterly_reserve': quarterly_reserve,
            'annual_reserve': annual_reserve,
            'sv_monthly': sv_monthly,
            'compensation': compensation,
            'monthly_fot': monthly_fot,
            'total_fot': total_fot,
            'months_count': months_count
        }

    def add_additional_position(self, position_data):
        """Добавить новую позицию с рассчитанным ФОТ"""
        fot_calculation = self.calculate_fot_for_position(position_data)
        position_entry = {
            'position_name': position_data.get('position_name', 'Новая позиция'),
            'parameters': position_data,
            'fot_calculation': fot_calculation
        }
        self.additional_positions.append(position_entry)
        self.changes_log.append({
            'action': 'Добавлена позиция',
            'fio': position_data.get('position_name', 'Новая позиция'),
            'month': 1,
            'original_index': -1,
            'prev_data': {},
            'new_data': position_data,
            'fot_total': fot_calculation['total_fot']
        })
        return fot_calculation

    def get_total_additional_fot(self):
        """Получить суммарный ФОТ всех дополнительных позиций"""
        return sum(pos['fot_calculation']['total_fot'] for pos in self.additional_positions)

    def get_employee_for_calculation(self, index):
        if index in self.modified_employees:
            entry = self.modified_employees[index]
            return entry['employee'] if isinstance(entry, dict) and 'employee' in entry else entry
        if 0 <= index < len(self.employees_data):
            return self.employees_data[index]
        return None

    def get_employee_for_month(self, index, month_index):
        if not (0 <= index < len(self.employees_data)):
            return None
        base = self.employees_data[index]
        if index not in self.modified_employees:
            return base
        entry = self.modified_employees[index]
        if isinstance(entry, dict) and 'employee' in entry:
            eff = max(1, min(12, int(entry.get('effective_from_month', 1))))
            emp = entry['employee']
        else:
            eff, emp = 1, entry
        return base if month_index + 1 < eff else emp

    def get_modification_effective_month(self, index):
        entry = self.modified_employees.get(index)
        if not entry:
            return 1
        if isinstance(entry, dict) and 'employee' in entry:
            return max(1, min(12, int(entry.get('effective_from_month', 1))))
        return 1

    def update_employee(self, index, updated_data, effective_from_month=1):
        eff = max(1, min(12, int(effective_from_month)))
        clean = dict(updated_data)
        clean.pop('effective_from_month', None)
        prev_emp = self.get_employee_for_calculation(index)
        prev_data = copy.deepcopy(prev_emp) if prev_emp else {}
        if 'base_salary' in clean:
            new_salary = float(clean['base_salary'] or 0)
            salary_arr = list(clean.get('base_salary_monthly',
                                        prev_data.get('base_salary_monthly', [0.0] * 12)))
            while len(salary_arr) < 12:
                salary_arr.append(0.0)
            for m in range(eff - 1, 12):
                salary_arr[m] = new_salary
            clean['base_salary_monthly'] = salary_arr
        self.modified_employees[index] = {'employee': clean, 'effective_from_month': eff}
        self.changes_log.append({
            'action': 'Изменен',
            'fio': clean.get('fio', 'Сотрудник'),
            'month': eff,
            'original_index': index,
            'prev_data': prev_data,
            'new_data': clean
        })
        return True

    def set_zero_headcount(self, index, month_index):
        emp = self.get_employee_for_calculation(index)
        if not emp:
            return False
        prev_data = copy.deepcopy(emp)
        headcounts = list(emp.get('headcounts', [1.0] * 12))
        if 0 <= month_index < len(headcounts):
            headcounts[month_index] = 0.0
        clean = copy.deepcopy(emp)
        clean['headcounts'] = headcounts
        clean.pop('original_index', None)
        self.update_employee(index, clean, month_index + 1)
        self.changes_log[-1] = {
            'action': 'Выведена позиция',
            'fio': clean.get('fio', ''),
            'month': month_index + 1,
            'original_index': index,
            'prev_data': prev_data,
            'new_data': clean,
            'details': f"Численность в {self.months[month_index]} -> 0 чел"
        }
        return True

    def reset_employee_changes(self, index):
        if index in self.modified_employees:
            del self.modified_employees[index]
            self.changes_log = [log for log in self.changes_log if log.get('original_index') != index]
            return True
        return False

    def search_employees(self, fio_query='', position_query=''):
        results = []
        if not self.employees_data:
            return results
        fio_q = fio_query.lower().strip() if fio_query else ''
        pos_q = position_query.lower().strip() if position_query else ''
        for idx, emp in enumerate(self.employees_data):
            if (fio_q and fio_q in emp['fio'].lower()) or (pos_q and pos_q in emp['position'].lower()):
                res = emp.copy()
                res['original_index'] = idx
                results.append(res)
        return results

    def calculate_monthly_fot_detailed(self, month_index, employee_index=None):
        """Расчёт ФОТ за конкретный месяц"""
        if employee_index is None:
            return None
        emp = self.get_employee_for_month(employee_index, month_index)
        if emp is None:
            return None
        comp_s = emp.get('compensation_summer', 0.0)
        comp_w = emp.get('compensation_winter', 0.0)
        sv_monthly_data = emp.get('excel_sv_monthly', [0.0] * 12)
        income_ytd, sv_ytd = 0.0, 0.0
        target = None
        base_salary_arr = emp.get('base_salary_monthly', [emp.get('base_salary', 0.0)] * 12)
        for m in range(1, 13):
            base_m = base_salary_arr[m - 1] if m - 1 < len(base_salary_arr) else 0.0
            hc = emp['headcounts'][m - 1] if m - 1 < len(emp['headcounts']) else 1.0
            if hc > 0:
                ms = base_m
                mb_pct = emp.get('monthly_bonus_percent', 0.0) / 100 if emp.get('monthly_bonus_percent',
                                                                                0.0) > 1 else emp.get(
                    'monthly_bonus_percent', 0.0)
                mb = ms * mb_pct
                income_ytd += ms + mb
                sv_m = sv_monthly_data[m - 1]
                sv_ytd += sv_m
            else:
                ms, mb, sv_m = 0.0, 0.0, 0.0
            comp_m = comp_w if m >= 10 or m <= 3 else comp_s
            res_q = emp['excel_res_q_monthly'][m - 1] if m - 1 < len(emp['excel_res_q_monthly']) else 0.0
            res_a = emp['excel_res_a_monthly'][m - 1] if m - 1 < len(emp['excel_res_a_monthly']) else 0.0
            fot_m = ms + mb + sv_m + res_q + res_a + comp_m
            if m == month_index + 1:
                target = {
                    'month_index': month_index, 'month_name': self.months[month_index],
                    'monthly_salary': ms, 'monthly_bonus': mb,
                    'cumulative_income': income_ytd, 'social_contributions': sv_m,
                    'compensations': comp_m, 'quarterly_reserve': res_q,
                    'annual_reserve': res_a, 'fot_total': fot_m
                }
        return target

    def calculate_total_fot_for_employee(self, index):
        """Расчёт общего ФОТ сотрудника за год"""
        emp = self.get_employee_for_calculation(index)
        if not emp:
            return 0.0
        total_fot = 0.0
        for m in range(12):
            detail = self.calculate_monthly_fot_detailed(m, employee_index=index)
            if detail:
                total_fot += detail['fot_total']
        total_fot += emp.get('gross_additions', 0.0)
        return total_fot

    def create_empty_employee(self):
        return {
            'directorate': '', 'fio': '', 'position': '', 'base_salary': 0.0,
            'base_salary_monthly': [0.0] * 12, 'allowance_cat1': 0.0, 'allowance_cat2': 0.0,
            'travel_allowance': 0.0, 'rk_percent': 0.0, 'sn_percent': 0.0,
            'harmfulness_percent': 0.0, 'night_percent': 0.0, 'equipment_compensation': 0.0,
            'car_rent': 0.0, 'crew_rent': 0.0, 'fuel_limit_summer': 0.0, 'fuel_limit_winter': 0.0,
            'transport_expenses': 0.0, 'car_compensation': 0.0, 'phone_limit': 0.0, 'internet_limit': 0.0,
            'dms': 0.0, 'housing_compensation': 0.0, 'travel_compensation': 0.0,
            'compensation_summer': 0.0, 'compensation_winter': 0.0,
            'monthly_bonus_percent': 0.0, 'quarterly_bonus_percent': 0.0, 'annual_bonus_percent': 0.0,
            'salary_type': '', 'last_employee': '', 'headcounts': [1.0] * 12,
            'excel_sv_monthly': [0.0] * 12, 'excel_sv_total': 0.0,
            'excel_comp_monthly': [0.0] * 12, 'excel_comp_total': 0.0,
            'excel_res_q_monthly': [0.0] * 12, 'excel_res_q_total': 0.0,
            'excel_res_a_monthly': [0.0] * 12, 'excel_res_a_total': 0.0,
            'base_fot_year': 0.0, 'gross_additions': 0.0, 'comment': ''
        }

    def add_employee(self, emp_data):
        new_index = len(self.employees_data)
        self.employees_data.append(emp_data)
        self.changes_log.append({
            'action': 'Добавлен сотрудник', 'fio': emp_data.get('fio', 'Без ФИО'),
            'month': 1, 'original_index': new_index, 'prev_data': {}, 'new_data': emp_data
        })
        return new_index


class ChangeLogDetailDialog:
    """Диалоговое окно для просмотра деталей записи из журнала изменений"""
    FIELD_NAMES = {
        'fio': 'ФИО', 'position': 'Должность', 'base_salary': 'Оклад (руб.)',
        'rk_percent': 'РК (%)', 'sn_percent': 'СН (%)',
        'monthly_bonus_percent': 'Месячная премия (%)',
        'quarterly_bonus_percent': 'Квартальная премия (%)',
        'annual_bonus_percent': 'Годовая премия (%)',
        'gross_additions': 'Доплаты ГРОСС (руб.)',
    }

    def __init__(self, parent, log_entry, calculator, gui=None):
        self.calculator = calculator
        self.gui = gui
        self.log_entry = log_entry
        self.dialog = tk.Toplevel(parent)
        self.dialog.title(f"Детали изменения: {log_entry.get('fio', '')}")
        self.dialog.geometry("680x620")
        self.dialog.transient(parent)
        self.dialog.grab_set()
        main_frame = ttk.Frame(self.dialog, padding=10)
        main_frame.pack(fill=tk.BOTH, expand=True)
        header = ttk.Label(main_frame, text=f"Сотрудник: {log_entry.get('fio', 'Неизвестно')}",
                           font=("Arial", 12, "bold"))
        header.pack(anchor="w", pady=(0, 5))
        action_text = log_entry.get('action', '')
        month = log_entry.get('month', 1)
        month_name = calculator.months[month - 1] if 1 <= month <= 12 else str(month)
        subheader = ttk.Label(main_frame,
                              text=f"Действие: {action_text} | Эффективно с: {month_name.capitalize()}",
                              font=("Arial", 10))
        subheader.pack(anchor="w", pady=(0, 10))
        self.text_widget = tk.Text(main_frame, wrap=tk.WORD, font=("Consolas", 10),
                                   bg="#f9f9f9", relief=tk.FLAT)
        self.text_widget.pack(fill=tk.BOTH, expand=True, pady=5)
        self.text_widget.tag_configure("was", foreground="#cc0000")
        self.text_widget.tag_configure("now", foreground="#007700")
        self.text_widget.tag_configure("comment_text", foreground="#0055aa", font=("Consolas", 10, "italic"))
        self.text_widget.tag_configure("section", foreground="#555555", font=("Consolas", 10, "bold"))
        self._generate_details(log_entry)
        self.text_widget.config(state=tk.DISABLED)
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill=tk.X, pady=(10, 0))
        # ✅ Кнопка "Отменить изменение"
        if action_text == 'Изменен' and log_entry.get('original_index') is not None:
            self.btn_undo = ttk.Button(btn_frame, text="Отменить изменение", command=self._undo_change)
            self.btn_undo.pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(btn_frame, text="Закрыть", command=self.dialog.destroy).pack(side=tk.RIGHT)

    def _undo_change(self):
        """Отменить изменение и восстановить оригинальные данные"""
        original_index = self.log_entry.get('original_index')
        if original_index is None:
            return
        if original_index >= len(self.calculator.original_employees_data):
            messagebox.showerror("Ошибка", "Оригинальные данные не найдены")
            return
        original_emp = self.calculator.original_employees_data[original_index]
        if original_index not in self.calculator.modified_employees:
            messagebox.showwarning("Внимание", "Изменения уже были отменены")
            return
        modified_entry = self.calculator.modified_employees[original_index]
        current_emp = modified_entry['employee']
        eff_month = modified_entry.get('effective_from_month', 1)
        changed_fields = []
        for key in self.FIELD_NAMES.keys():
            old_val = self.log_entry.get('prev_data', {}).get(key)
            new_val = self.log_entry.get('new_data', {}).get(key)
            if old_val != new_val:
                changed_fields.append(key)
        old_hc = self.log_entry.get('prev_data', {}).get('headcounts', [])
        new_hc = self.log_entry.get('new_data', {}).get('headcounts', [])
        hc_changed = old_hc != new_hc
        restored_emp = copy.deepcopy(current_emp)
        for field in changed_fields:
            if field in original_emp:
                restored_emp[field] = original_emp[field]
        if hc_changed:
            restored_emp['headcounts'] = list(original_emp.get('headcounts', [1.0] * 12))
        if 'base_salary' in changed_fields or 'quarterly_bonus_percent' in changed_fields:
            restored_emp['base_salary_monthly'] = list(
                original_emp.get('base_salary_monthly', [original_emp.get('base_salary', 0.0)] * 12))
            orig_sv = list(original_emp.get('excel_sv_monthly', [0.0] * 12))
            orig_res_q = list(original_emp.get('excel_res_q_monthly', [0.0] * 12))
            orig_res_a = list(original_emp.get('excel_res_a_monthly', [0.0] * 12))
            restored_emp['excel_sv_monthly'] = orig_sv
            restored_emp['excel_sv_total'] = sum(orig_sv)
            restored_emp['excel_res_q_monthly'] = orig_res_q
            restored_emp['excel_res_q_total'] = sum(orig_res_q)
            restored_emp['excel_res_a_monthly'] = orig_res_a
            restored_emp['excel_res_a_total'] = sum(orig_res_a)
        self.calculator.modified_employees[original_index] = {
            'employee': restored_emp,
            'effective_from_month': eff_month
        }
        self.calculator.changes_log.remove(self.log_entry)
        messagebox.showinfo("Успех",
                            f"Изменение для '{self.log_entry.get('fio', '')}' отменено.\nВосстановлено полей: {len(changed_fields)}")
        self.dialog.destroy()
        if self.gui:
            self.gui._refresh_history()
            self.gui._update_results()

    @staticmethod
    def _fmt_percent(val):
        if val is None or val == '': return '0'
        try:
            v = float(val)
            display = v * 100 if v <= 1 else v
            s = f"{display:.4f}".rstrip('0').rstrip('.')
            return s if s else "0"
        except:
            return str(val)

    @staticmethod
    def _fmt_money(val):
        if val is None or val == '': return '0'
        try:
            return f"{float(val):,.2f}".replace(",", " ")
        except:
            return str(val)

    @staticmethod
    def _fmt_headcount(val):
        if val is None: return '0'
        try:
            v = float(val)
            return str(int(v)) if v == int(v) else str(v)
        except:
            return str(val)

    def _generate_details(self, log_entry):
        self.text_widget.config(state=tk.NORMAL)
        self.text_widget.delete(1.0, tk.END)
        action = log_entry.get('action', '')
        new_data = log_entry.get('new_data', {})
        prev_data = log_entry.get('prev_data', {})

        # ✅ ФУНКЦИЯ ДЛЯ РАСЧЕТА ФОТ СОТРУДНИКА
        def calculate_employee_fot(emp_data):
            """Расчет годового ФОТ для сотрудника"""
            if not emp_data:
                return 0.0

            # Создаем временную копию с данными сотрудника
            temp_emp = self.calculator.create_empty_employee()
            temp_emp.update(emp_data)

            # Убеждаемся, что все необходимые поля есть
            if 'headcounts' not in temp_emp:
                temp_emp['headcounts'] = [1.0] * 12
            if 'base_salary_monthly' not in temp_emp:
                base_sal = temp_emp.get('base_salary', 0.0)
                temp_emp['base_salary_monthly'] = [base_sal] * 12
            if 'excel_sv_monthly' not in temp_emp:
                temp_emp['excel_sv_monthly'] = [0.0] * 12
            if 'excel_res_q_monthly' not in temp_emp:
                temp_emp['excel_res_q_monthly'] = [0.0] * 12
            if 'excel_res_a_monthly' not in temp_emp:
                temp_emp['excel_res_a_monthly'] = [0.0] * 12

            total_fot = 0.0
            for m in range(12):
                base_m = temp_emp['base_salary_monthly'][m] if m < len(temp_emp['base_salary_monthly']) else 0.0
                hc = temp_emp['headcounts'][m] if m < len(temp_emp['headcounts']) else 1.0

                if hc > 0:
                    ms = base_m
                    mb_pct = temp_emp.get('monthly_bonus_percent', 0.0)
                    if mb_pct > 1:
                        mb_pct = mb_pct / 100.0
                    mb = ms * mb_pct

                    sv_m = temp_emp['excel_sv_monthly'][m] if m < len(temp_emp['excel_sv_monthly']) else 0.0
                    res_q = temp_emp['excel_res_q_monthly'][m] if m < len(temp_emp['excel_res_q_monthly']) else 0.0
                    res_a = temp_emp['excel_res_a_monthly'][m] if m < len(temp_emp['excel_res_a_monthly']) else 0.0

                    comp_s = temp_emp.get('compensation_summer', 0.0)
                    comp_w = temp_emp.get('compensation_winter', 0.0)
                    comp_m = comp_w if m >= 9 or m <= 2 else comp_s

                    fot_m = ms + mb + sv_m + res_q + res_a + comp_m
                    total_fot += fot_m

            # Добавляем доплаты gross
            total_fot += temp_emp.get('gross_additions', 0.0)
            return total_fot

        if action == 'Добавлена позиция':
            self.text_widget.insert(tk.END, "Добавленная позиция:\n", "section")
            self.text_widget.insert(tk.END, "=" * 50 + "\n\n")
            fot_calc = log_entry.get('fot_total', 0)
            self.text_widget.insert(tk.END, f"  Итого ФОТ: ", "section")
            self.text_widget.insert(tk.END, f"{fot_calc:,.2f} руб.\n\n")
            self.text_widget.insert(tk.END, "Параметры:\n", "section")
            for key, label in self.FIELD_NAMES.items():
                val = new_data.get(key)
                if val is not None and val != 0:
                    if key.endswith('_percent'):
                        val_str = self._fmt_percent(val)
                    elif key == 'base_salary':
                        val_str = self._fmt_money(val)
                    else:
                        val_str = str(val)
                    self.text_widget.insert(tk.END, f"  {label}: {val_str}\n")
            self.text_widget.config(state=tk.DISABLED)
            return

        if action in ('Изменен', 'Добавлен сотрудник'):
            self.text_widget.insert(tk.END, "Изменённые параметры:\n", "section")
            self.text_widget.insert(tk.END, "=" * 50 + "\n\n")
            changes_found = False
            for key, label in self.FIELD_NAMES.items():
                old_val = prev_data.get(key)
                new_val = new_data.get(key)
                if key.endswith('_percent'):
                    old_str = self._fmt_percent(old_val)
                    new_str = self._fmt_percent(new_val)
                elif key in ('base_salary', 'gross_additions'):
                    old_str = self._fmt_money(old_val)
                    new_str = self._fmt_money(new_val)
                else:
                    old_str = str(old_val).strip() if old_val is not None else ''
                    new_str = str(new_val).strip() if new_val is not None else ''
                if old_str != new_str:
                    changes_found = True
                    self.text_widget.insert(tk.END, f"  {label}:\n")
                    self.text_widget.insert(tk.END, f"    Было: ", "was")
                    self.text_widget.insert(tk.END, f"{old_str if old_str else '—'}\n")
                    self.text_widget.insert(tk.END, f"    Стало: ", "now")
                    self.text_widget.insert(tk.END, f"{new_str if new_str else '—'}\n\n")

            old_hc = prev_data.get('headcounts', [])
            new_hc = new_data.get('headcounts', [])
            if old_hc and new_hc:
                hc_changes = []
                for i in range(12):
                    old_v = old_hc[i] if i < len(old_hc) else 1.0
                    new_v = new_hc[i] if i < len(new_hc) else 1.0
                    if float(old_v) != float(new_v):
                        hc_changes.append((i, old_v, new_v))
                if hc_changes:
                    changes_found = True
                    self.text_widget.insert(tk.END, "  Численность:\n")
                    for m_idx, ov, nv in hc_changes:
                        m_name = self.calculator.months[m_idx].capitalize()
                        self.text_widget.insert(tk.END, f"    {m_name}: ")
                        self.text_widget.insert(tk.END, f"Было: {self._fmt_headcount(ov)}", "was")
                        self.text_widget.insert(tk.END, " → ")
                        self.text_widget.insert(tk.END, f"Стало: {self._fmt_headcount(nv)}\n", "now")
                    self.text_widget.insert(tk.END, "\n")

            comment_val = new_data.get('comment', '')
            if comment_val and str(comment_val).strip():
                changes_found = True
                self.text_widget.insert(tk.END, f"  Комментарий: ", "section")
                self.text_widget.insert(tk.END, f"\"{str(comment_val).strip()}\"\n", "comment_text")

            if not changes_found:
                self.text_widget.insert(tk.END, "  Существенные изменения не обнаружены.\n")

            # ✅ ОТОБРАЖЕНИЕ ФОТ ДО И ПОСЛЕ ИЗМЕНЕНИЙ
            self.text_widget.insert(tk.END, "\n" + "=" * 50 + "\n")
            self.text_widget.insert(tk.END, "РАСЧЕТ ФОТ\n", "section")
            self.text_widget.insert(tk.END, "=" * 50 + "\n\n")

            fot_before = calculate_employee_fot(prev_data)
            fot_after = calculate_employee_fot(new_data)
            fot_difference = fot_after - fot_before

            self.text_widget.insert(tk.END, f"  ФОТ до изменений:  ", "section")
            self.text_widget.insert(tk.END, f"{fot_before:,.2f} руб.\n", "was")
            self.text_widget.insert(tk.END, f"  ФОТ после изменений: ", "section")
            self.text_widget.insert(tk.END, f"{fot_after:,.2f} руб.\n", "now")
            self.text_widget.insert(tk.END, f"  Разница:            ", "section")

            if fot_difference > 0:
                self.text_widget.insert(tk.END, f"+{fot_difference:,.2f} руб. (увеличение)\n", "now")
            elif fot_difference < 0:
                self.text_widget.insert(tk.END, f"{fot_difference:,.2f} руб. (уменьшение)\n", "was")
            else:
                self.text_widget.insert(tk.END, f"0.00 руб. (без изменений)\n")

        elif action == 'Выведена позиция':
            self.text_widget.insert(tk.END, "Детали операции:\n", "section")
            self.text_widget.insert(tk.END, "=" * 50 + "\n\n")
            self.text_widget.insert(tk.END, f"  {log_entry.get('details', 'Численность обнулена')}\n")
            old_hc = prev_data.get('headcounts', [])
            new_hc = new_data.get('headcounts', [])
            if old_hc and new_hc:
                for i in range(12):
                    old_v = old_hc[i] if i < len(old_hc) else 1.0
                    new_v = new_hc[i] if i < len(new_hc) else 1.0
                    if float(old_v) != float(new_v):
                        m_name = self.calculator.months[i].capitalize()
                        self.text_widget.insert(tk.END, f"  {m_name}: ")
                        self.text_widget.insert(tk.END, f"Было: {self._fmt_headcount(old_v)}", "was")
                        self.text_widget.insert(tk.END, " → ")
                        self.text_widget.insert(tk.END, f"Стало: {self._fmt_headcount(new_v)}\n", "now")

            # ✅ ОТОБРАЖЕНИЕ ФОТ ДЛЯ ВЫВЕДЕННОЙ ПОЗИЦИИ
            self.text_widget.insert(tk.END, "\n" + "=" * 50 + "\n")
            self.text_widget.insert(tk.END, "РАСЧЕТ ФОТ\n", "section")
            self.text_widget.insert(tk.END, "=" * 50 + "\n\n")

            fot_before = calculate_employee_fot(prev_data)
            fot_after = calculate_employee_fot(new_data)
            fot_difference = fot_after - fot_before

            self.text_widget.insert(tk.END, f"  ФОТ до выведения:  ", "section")
            self.text_widget.insert(tk.END, f"{fot_before:,.2f} руб.\n", "was")
            self.text_widget.insert(tk.END, f"  ФОТ после выведения: ", "section")
            self.text_widget.insert(tk.END, f"{fot_after:,.2f} руб.\n", "now")
            self.text_widget.insert(tk.END, f"  Экономия:           ", "section")
            self.text_widget.insert(tk.END, f"{abs(fot_difference):,.2f} руб.\n", "was")

        self.text_widget.config(state=tk.DISABLED)


class AdditionalPositionDialog:
    """Диалог для добавления новой позиции с расчетом ФОТ"""

    def __init__(self, parent, calculator):
        self.calculator = calculator
        self.dialog = tk.Toplevel(parent)
        self.dialog.title("Добавить новую позицию")
        self.dialog.geometry("420x520")
        self.dialog.transient(parent)
        self.dialog.grab_set()
        self.entries = {}
        self.fot_calculation = None
        self._create_widgets()
        self.dialog.wait_window()

    def _create_widgets(self):
        main_frame = ttk.Frame(self.dialog, padding=15)
        main_frame.pack(fill=tk.BOTH, expand=True)
        ttk.Label(main_frame, text="Параметры новой позиции",
                  font=("Arial", 11, "bold")).pack(pady=(0, 15))
        fields = [
            ('Название позиции:', 'position_name', ''),
            ('Оклад (руб.):', 'base_salary', '0'),
            ('РК (%):', 'rk_percent', '0'),
            ('СН (%):', 'sn_percent', '0'),
            ('Месячная премия (%):', 'monthly_bonus_percent', '0'),
            ('Квартальная премия (%):', 'quarterly_bonus_percent', '0'),
            ('Годовая премия (%):', 'annual_bonus_percent', '0'),
            ('Компенсации лето (руб.):', 'compensation_summer', '0'),
            ('Компенсации зима (руб.):', 'compensation_winter', '0'),
            ('Количество месяцев:', 'months_count', '12'),
        ]
        for label, key, default in fields:
            frame = ttk.Frame(main_frame)
            frame.pack(fill=tk.X, pady=2)
            ttk.Label(frame, text=label, width=28, anchor='e').pack(side=tk.LEFT)
            entry = ttk.Entry(frame, width=12, justify='right')
            entry.insert(0, default)
            entry.pack(side=tk.RIGHT, padx=(5, 0))
            self.entries[key] = entry
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill=tk.X, pady=(15, 0))
        self.btn_calculate = ttk.Button(btn_frame, text="Рассчитать",
                                        command=self._calculate_fot)
        self.btn_calculate.pack(side=tk.LEFT, padx=(0, 10))
        self.btn_add = ttk.Button(btn_frame, text="Добавить",
                                  command=self._add_position, state=tk.DISABLED)
        self.btn_add.pack(side=tk.LEFT)
        ttk.Button(btn_frame, text="Отмена",
                   command=self.dialog.destroy).pack(side=tk.RIGHT, padx=(10, 0))
        self.result_frame = ttk.LabelFrame(main_frame, text="Результат расчета ФОТ", padding=10)
        self.result_frame.pack(fill=tk.X, pady=(15, 0))
        self.result_frame.pack_forget()
        self.result_labels = {}

    def _calculate_fot(self):
        try:
            position_data = {}
            for key, entry in self.entries.items():
                value = entry.get().strip()
                if key in ['position_name']:
                    position_data[key] = value
                else:
                    position_data[key] = self.calculator._parse_number(value)
            if not position_data.get('position_name'):
                messagebox.showerror("Ошибка", "Введите название позиции")
                return
            if position_data.get('base_salary', 0) <= 0:
                messagebox.showerror("Ошибка", "Оклад должен быть больше 0")
                return
            self.fot_calculation = self.calculator.calculate_fot_for_position(position_data)
            self._show_results(position_data, self.fot_calculation)
            self.btn_add.config(state=tk.NORMAL)
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка при расчете: {str(e)}")

    def _show_results(self, position_data, fot_calc):
        self.result_frame.pack(fill=tk.X, pady=(15, 0))
        for widget in self.result_frame.winfo_children():
            widget.destroy()
        self.result_labels.clear()

        def fmt(val):
            return f"{val:,.2f}".replace(",", " ")

        results = [
            ("Название:", position_data['position_name']),
            ("ФОТ за месяц:", f"{fmt(fot_calc['monthly_fot'])} руб."),
            ("ФОТ за период:", f"{fmt(fot_calc['total_fot'])} руб."),
            ("", ""),
        ]
        for i, (label, value) in enumerate(results):
            if label:
                lbl = ttk.Label(self.result_frame, text=f"{label:<25} {value}",
                                font=("Arial", 9 if i < 4 else 8))
                lbl.pack(anchor='w', pady=1)

    def _add_position(self):
        if not self.fot_calculation:
            return
        position_data = {}
        for key, entry in self.entries.items():
            value = entry.get().strip()
            if key in ['position_name']:
                position_data[key] = value
            else:
                position_data[key] = self.calculator._parse_number(value)
        self.calculator.add_additional_position(position_data)
        messagebox.showinfo("Успех",
                            f"Позиция '{position_data['position_name']}' добавлена!\n"
                            f"ФОТ за {self.fot_calculation['months_count']} мес.: "
                            f"{self.fot_calculation['total_fot']:,.2f} руб.")
        self.dialog.destroy()


class HeadcountEditDialog:
    """Диалог для изменения численности и доп. выплат"""

    def __init__(self, parent, calculator, employee_index, month_index, gui=None):
        self.calculator = calculator
        self.employee_index = employee_index
        self.month_index = month_index
        self.gui = gui
        self.result = None

        self.dialog = tk.Toplevel(parent)
        self.dialog.title("Изменение численности и доп. выплат")
        self.dialog.geometry("450x280")
        self.dialog.transient(parent)
        self.dialog.grab_set()

        self._create_widgets()
        self.dialog.wait_window()

    def _create_widgets(self):
        main_frame = ttk.Frame(self.dialog, padding=15)
        main_frame.pack(fill=tk.BOTH, expand=True)

        # Получаем текущие значения
        emp = self.calculator.get_employee_for_calculation(self.employee_index)
        current_headcount = emp['headcounts'][self.month_index] if emp and 'headcounts' in emp else 1.0
        current_gross = emp.get('gross_additions', 0.0) if emp else 0.0

        # Поле численности
        hc_frame = ttk.Frame(main_frame)
        hc_frame.pack(fill=tk.X, pady=5)
        ttk.Label(hc_frame, text="Численность:", width=20, anchor='e').pack(side=tk.LEFT)
        self.headcount_var = tk.DoubleVar(value=current_headcount)
        self.headcount_spinbox = ttk.Spinbox(
            hc_frame,
            from_=0.0,
            to=1.0,
            increment=0.1,
            textvariable=self.headcount_var,
            width=10
        )
        self.headcount_spinbox.pack(side=tk.RIGHT, padx=(5, 0))

        # Поле доп. выплат
        gross_frame = ttk.Frame(main_frame)
        gross_frame.pack(fill=tk.X, pady=5)
        ttk.Label(gross_frame, text="Доп. выплаты (руб.):", width=20, anchor='e').pack(side=tk.LEFT)
        self.gross_entry = ttk.Entry(gross_frame, width=12, justify='right')
        self.gross_entry.insert(0, str(current_gross))
        self.gross_entry.pack(side=tk.RIGHT, padx=(5, 0))

        # Выбор первого месяца
        start_month_frame = ttk.Frame(main_frame)
        start_month_frame.pack(fill=tk.X, pady=5)
        ttk.Label(start_month_frame, text="Первый месяц:", width=20, anchor='e').pack(side=tk.LEFT)
        self.start_month_combo = ttk.Combobox(
            start_month_frame,
            values=self.calculator.months,
            state="readonly",
            width=27
        )
        self.start_month_combo.current(self.month_index)
        self.start_month_combo.pack(side=tk.RIGHT, padx=(5, 0))

        # Выбор последнего месяца
        end_month_frame = ttk.Frame(main_frame)
        end_month_frame.pack(fill=tk.X, pady=5)
        ttk.Label(end_month_frame, text="Последний месяц:", width=20, anchor='e').pack(side=tk.LEFT)
        self.end_month_combo = ttk.Combobox(
            end_month_frame,
            values=self.calculator.months,
            state="readonly",
            width=27
        )
        self.end_month_combo.current(self.month_index)
        self.end_month_combo.pack(side=tk.RIGHT, padx=(5, 0))

        # Кнопки
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill=tk.X, pady=(15, 0))
        ttk.Button(btn_frame, text="Применить", command=self._apply).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(btn_frame, text="Отмена", command=self.dialog.destroy).pack(side=tk.RIGHT)

    def _apply(self):
        try:
            new_headcount = self.headcount_var.get()
            new_gross = self.calculator._parse_number(self.gross_entry.get())

            if new_headcount < 0 or new_headcount > 1:
                messagebox.showerror("Ошибка", "Численность должна быть от 0 до 1")
                return

            # Получаем выбранные месяцы
            start_month = self.start_month_combo.current()
            end_month = self.end_month_combo.current()

            if start_month > end_month:
                messagebox.showerror("Ошибка", "Первый месяц не может быть позже последнего")
                return

            # Получаем текущего сотрудника
            emp = self.calculator.get_employee_for_calculation(self.employee_index)
            if not emp:
                messagebox.showerror("Ошибка", "Сотрудник не найден")
                return

            # Создаем копию для изменений
            updated_emp = copy.deepcopy(emp)

            # Обновляем численность только в выбранных месяцах
            headcounts = list(updated_emp.get('headcounts', [1.0] * 12))
            for m in range(start_month, end_month + 1):
                headcounts[m] = new_headcount
            updated_emp['headcounts'] = headcounts

            # Обновляем доп. выплаты
            updated_emp['gross_additions'] = new_gross

            # Сохраняем изменения
            self.calculator.update_employee(self.employee_index, updated_emp, start_month + 1)

            # Обновляем запись в логе
            if self.calculator.changes_log:
                last_log = self.calculator.changes_log[-1]
                if last_log.get('original_index') == self.employee_index:
                    last_log['new_data'] = updated_emp
                    if new_headcount == 0:
                        last_log['action'] = 'Выведена позиция'
                        start_name = self.calculator.months[start_month]
                        end_name = self.calculator.months[end_month]
                        if start_month == end_month:
                            last_log['details'] = f"Численность в {start_name} -> {new_headcount} чел"
                        else:
                            last_log['details'] = f"Численность с {start_name} по {end_name} -> {new_headcount} чел"
                    else:
                        last_log['action'] = 'Изменен'

            self.result = {
                'headcount': new_headcount,
                'gross_additions': new_gross,
                'start_month': start_month,
                'end_month': end_month
            }

            messagebox.showinfo("Успех", f"Изменения применены для '{emp.get('fio', '')}'\n"
                                         f"Период: {self.calculator.months[start_month]} - {self.calculator.months[end_month]}")
            self.dialog.destroy()

            if self.gui:
                self.gui._refresh_history()
                self.gui._update_results()

        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка при применении изменений: {str(e)}")

class EmployeeEditDialog:
    """Диалог редактирования"""
    _PERCENT_FIELD_LABELS = {
        'rk_percent': 'РК (%)', 'sn_percent': 'СН (%)',
        'monthly_bonus_percent': 'Месячная премия (%)',
        'quarterly_bonus_percent': 'Квартальная премия (%)',
        'annual_bonus_percent': 'Годовая премия (%)',
    }

    def __init__(self, parent, calculator):
        self.calculator = calculator
        self.dialog = tk.Toplevel(parent)
        self.dialog.title("Редактирование данных")
        self.dialog.geometry("950x850")
        self.dialog.transient(parent)
        self.dialog.grab_set()
        self.selected_employee = None
        self.original_index = None
        self.current_employee = None
        self.entries = {}
        self.edit_widgets = []
        self.effective_month_combo = None
        self.is_adding_new = False
        self._create_widgets()
        self.dialog.protocol("WM_DELETE_WINDOW", self._on_close)
        self.dialog.wait_window()

    def _on_close(self):
        try:
            if hasattr(self, 'canvas') and self.canvas.winfo_exists():
                self.canvas.unbind_all("<MouseWheel>")
        except:
            pass
        self.dialog.destroy()

    @staticmethod
    def _format_salary_display(value):
        try:
            return f"{float(value):.0f}"
        except:
            return "0"

    @staticmethod
    def _format_percent_display(value):
        try:
            v = float(value)
        except:
            return "0"
        if abs(v - round(v)) < 1e-9: return str(int(round(v)))
        s = f"{v:.4f}".rstrip('0').rstrip('.')
        return s if s else "0"

    @staticmethod
    def _format_text_display(value):
        return "  " if value is None else str(value).strip()

    def _on_mousewheel(self, event):
        try:
            if hasattr(self, 'canvas') and self.canvas.winfo_exists():
                if event.delta:
                    self.canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
                elif event.num == 4:
                    self.canvas.yview_scroll(-1, "units")
                elif event.num == 5:
                    self.canvas.yview_scroll(1, "units")
        except:
            pass

    def _create_widgets(self):
        mf = ttk.Frame(self.dialog, padding="10")
        mf.pack(fill=tk.BOTH, expand=True)
        self.canvas = tk.Canvas(mf, highlightthickness=0)
        self.scrollbar = ttk.Scrollbar(mf, orient="vertical", command=self.canvas.yview)
        self.scrollable_frame = ttk.Frame(self.canvas)
        self.scrollable_frame.bind("<Configure>", lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all")))
        self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        self.canvas.configure(yscrollcommand=self.scrollbar.set)
        self.canvas.pack(side="left", fill="both", expand=True)
        self.scrollbar.pack(side="right", fill="y")
        self.canvas.bind_all("<MouseWheel>", self._on_mousewheel)
        sf = ttk.LabelFrame(self.scrollable_frame, text="Поиск сотрудника", padding="10")
        sf.pack(fill=tk.X, pady=(0, 10))
        ttk.Label(sf, text="ФИО:   ").grid(row=0, column=0, sticky="w", padx=(0, 10))
        self.search_fio_entry = ttk.Entry(sf, width=40)
        self.search_fio_entry.grid(row=0, column=1, sticky="ew", padx=(0, 20))
        ttk.Label(sf, text="Должность:   ").grid(row=1, column=0, sticky="w", padx=(0, 10), pady=(10, 0))
        self.search_position_entry = ttk.Entry(sf, width=40)
        self.search_position_entry.grid(row=1, column=1, sticky="ew", padx=(0, 20), pady=(10, 0))
        btn_f = ttk.Frame(sf)
        btn_f.grid(row=2, column=0, columnspan=2, pady=(10, 0))
        ttk.Button(btn_f, text="Найти", command=self._search_employees).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(btn_f, text="Очистить", command=self._clear_search).pack(side=tk.LEFT)
        rf = ttk.LabelFrame(self.scrollable_frame, text="Результаты поиска", padding="10")
        rf.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        cols = ('fio', 'position', 'directorate', 'salary', 'status')
        self.tree = ttk.Treeview(rf, columns=cols, show='headings', height=8)
        for c, w, t in zip(cols, [250, 200, 150, 120, 100], ('ФИО', 'Должность', 'Дирекция', 'Оклад', 'Статус')):
            self.tree.heading(c, text=t)
            self.tree.column(c, width=w, anchor='e' if c == 'salary' else 'center' if c == 'status' else 'w')
        self.tree.pack(fill=tk.BOTH, expand=True)
        self.item_to_employee = {}
        self.tree.bind('<Double-1>', self._select_employee)
        ef = ttk.LabelFrame(self.scrollable_frame, text="Редактирование данных", padding="10")
        ef.pack(fill=tk.X, pady=(0, 10))
        self.edit_fields_frame = ttk.Frame(ef)
        self.edit_fields_frame.pack(fill=tk.X)
        self.no_sel_label = ttk.Label(self.edit_fields_frame, text="Выберите сотрудника из таблицы")
        self.no_sel_label.pack(pady=20)
        af = ttk.Frame(self.scrollable_frame)
        af.pack(fill=tk.X)
        self.btn_save = ttk.Button(af, text="Сохранить изменения", command=self._save_changes, state=tk.DISABLED)
        self.btn_save.pack(side=tk.LEFT, padx=(0, 10))
        self.btn_reset = ttk.Button(af, text="Сбросить изменения", command=self._reset_changes, state=tk.DISABLED)
        self.btn_reset.pack(side=tk.LEFT, padx=(0, 10))
        self.btn_zero = ttk.Button(af, text="Вывести позицию", command=self._set_zero_headcount,
                                   state=tk.DISABLED)
        self.btn_zero.pack(side=tk.LEFT, padx=(0, 10))
        self.btn_add = ttk.Button(af, text="Добавить позицию", command=self._add_position)
        self.btn_add.pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(af, text="Закрыть", command=self._on_close).pack(side=tk.RIGHT)

    def _modification_status_label(self, idx):
        if idx < 0 or idx not in self.calculator.modified_employees: return "Оригинал"
        m = self.calculator.get_modification_effective_month(idx)
        return f"Изменен (с {self.calculator.months[m - 1]})"

    def _current_employee_row_for_edit(self, emp):
        idx = emp.get('original_index', -1)
        if idx < 0: return emp
        row = self.calculator.get_employee_for_calculation(idx)
        return emp if row is None else {**row, 'original_index': idx}

    def _search_employees(self):
        fio = self.search_fio_entry.get().strip()
        pos = self.search_position_entry.get().strip()
        if not fio and not pos: return messagebox.showwarning("Предупреждение", "Введите хотя бы один параметр")
        if not self.calculator.employees_data: return messagebox.showwarning("Предупреждение",
                                                                             "Загрузите файл перед поиском")
        res = self.calculator.search_employees(fio, pos)
        for i in self.tree.get_children(): self.tree.delete(i)
        self.item_to_employee.clear()
        if not res: return messagebox.showinfo("Результат", "Сотрудники не найдены")
        for emp in res:
            idx = emp.get('original_index', -1)
            self.tree.insert('', tk.END, values=(emp['fio'], emp['position'], emp['directorate'],
                                                 self._format_salary_display(emp['base_salary']),
                                                 self._modification_status_label(idx)))
            self.item_to_employee[self.tree.get_children()[-1]] = emp

    def _clear_search(self):
        self.search_fio_entry.delete(0, tk.END)
        self.search_position_entry.delete(0, tk.END)
        for i in self.tree.get_children(): self.tree.delete(i)
        self.item_to_employee.clear()
        self._clear_edit_form()

    def _select_employee(self, event=None):
        sel = self.tree.selection()
        if not sel: return
        emp = self.item_to_employee.get(sel[0])
        if not emp: return
        self.selected_employee = self._current_employee_row_for_edit(emp)
        self.original_index = self.selected_employee.get('original_index', -1)
        self._show_edit_form(self.selected_employee)
        self.btn_save.config(state=tk.NORMAL)
        self.btn_reset.config(state=tk.NORMAL)
        self.btn_zero.config(state=tk.NORMAL)
        self.btn_add.config(state=tk.DISABLED)

    def _on_month_change(self, event=None):
        if not self.current_employee or 'base_salary' not in self.entries: return
        try:
            month_name = self.effective_month_combo.get()
            month_idx = self.calculator.months.index(month_name)
            salary_val = self.current_employee['base_salary_monthly'][month_idx]
            self.entries['base_salary'].delete(0, tk.END)
            self.entries['base_salary'].insert(0, self._format_salary_display(salary_val))
        except Exception:
            pass

    def _show_edit_form(self, emp):
        self._clear_edit_form()
        self.no_sel_label.pack_forget()
        self.current_employee = emp
        eff = self.calculator.get_modification_effective_month(
            self.original_index) if self.original_index is not None and self.original_index >= 0 else 1
        mr = ttk.Frame(self.edit_fields_frame)
        mr.pack(fill=tk.X, pady=(0, 8))
        ttk.Label(mr, text="Изменения с месяца:   ", width=25, anchor="w").pack(side=tk.LEFT, padx=(0, 10))
        self.effective_month_combo = ttk.Combobox(mr, width=22, state="readonly", values=self.calculator.months)
        self.effective_month_combo.current(eff - 1)
        self.effective_month_combo.pack(side=tk.LEFT, fill=tk.X, expand=True)
        self.edit_widgets.append(mr)
        self.effective_month_combo.bind("<<ComboboxSelected>>", self._on_month_change)
        for lbl, key in [("ФИО:   ", 'fio'), ("Должность:   ", 'position'), ("Оклад (руб.):   ", 'base_salary'),
                         ("РК (%):   ", 'rk_percent'), ("СН (%):   ", 'sn_percent'),
                         ("Месячная премия (%):   ", 'monthly_bonus_percent'),
                         ("Квартальная премия (%):   ", 'quarterly_bonus_percent'),
                         ("Годовая премия (%):   ", 'annual_bonus_percent')]:
            f = ttk.Frame(self.edit_fields_frame)
            f.pack(fill=tk.X, pady=5)
            ttk.Label(f, text=lbl, width=25, anchor="w").pack(side=tk.LEFT, padx=(0, 10))
            e = ttk.Entry(f, width=35)
            e.pack(side=tk.LEFT, fill=tk.X, expand=True)
            val = emp.get(key, '' if key in ('fio', 'position') else 0)
            if key == 'base_salary':
                month_idx = eff - 1
                if 'base_salary_monthly' in emp and len(emp['base_salary_monthly']) > month_idx:
                    val = emp['base_salary_monthly'][month_idx]
                else:
                    val = emp.get('base_salary', 0)
                e.insert(0, self._format_salary_display(val))
            elif key in self._PERCENT_FIELD_LABELS:
                e.insert(0, self._format_percent_display(val * 100 if val < 1 else val))
            else:
                e.insert(0, self._format_text_display(val))
            self.entries[key] = e
            self.edit_widgets.append(f)
        f_gross = ttk.Frame(self.edit_fields_frame)
        f_gross.pack(fill=tk.X, pady=5)
        ttk.Label(f_gross, text="Доплаты ГРОСС (руб.):   ", width=25, anchor="w").pack(side=tk.LEFT, padx=(0, 10))
        e_gross = ttk.Entry(f_gross, width=35)
        e_gross.pack(side=tk.LEFT, fill=tk.X, expand=True)
        e_gross.insert(0, self._format_salary_display(emp.get('gross_additions', 0)))
        self.entries['gross_additions'] = e_gross
        self.edit_widgets.append(f_gross)
        f_comm = ttk.Frame(self.edit_fields_frame)
        f_comm.pack(fill=tk.X, pady=5)
        ttk.Label(f_comm, text="Комментарий:   ", width=25, anchor="w").pack(side=tk.LEFT, padx=(0, 10))
        e_comm = ttk.Entry(f_comm, width=35)
        e_comm.pack(side=tk.LEFT, fill=tk.X, expand=True)
        e_comm.insert(0, emp.get('comment', ''))
        self.entries['comment'] = e_comm
        self.edit_widgets.append(f_comm)

    def _clear_edit_form(self):
        for w in self.edit_widgets: w.destroy()
        self.edit_widgets.clear()
        self.entries.clear()
        self.effective_month_combo = None
        self.no_sel_label.pack(pady=20)
        self.btn_save.config(state=tk.DISABLED)
        self.btn_reset.config(state=tk.DISABLED)
        if hasattr(self, 'btn_zero'): self.btn_zero.config(state=tk.DISABLED)

    def _save_changes(self):
        if self.original_index is None and not self.is_adding_new: return
        if self.is_adding_new:
            upd = self.calculator.create_empty_employee()
        else:
            upd = copy.deepcopy(self.selected_employee)
        for k, e in self.entries.items():
            v = e.get().strip()
            if k in ['base_salary', 'rk_percent', 'sn_percent', 'monthly_bonus_percent',
                     'quarterly_bonus_percent', 'annual_bonus_percent', 'gross_additions']:
                num = self.calculator._parse_number(v) if v else 0
                if k not in ('base_salary', 'gross_additions') and num > 100: num /= 100
                upd[k] = num
            else:
                upd[k] = v
        if self.is_adding_new:
            if not upd.get('fio', '').strip(): return messagebox.showwarning("Внимание", "Введите ФИО сотрудника")
            self.calculator.add_employee(upd)
            messagebox.showinfo("Успех", f"Сотрудник {upd.get('fio', '')} добавлен.")
        else:
            try:
                eff = self.calculator.months.index(self.effective_month_combo.get()) + 1
            except:
                eff = 1
            self.calculator.update_employee(self.original_index, upd, eff)
            messagebox.showinfo("Успех", f"Изменения сохранены: {upd.get('fio', '')}")
        self.dialog.destroy()

    def _add_position(self):
        """Открыть диалог для добавления новой позиции с расчетом ФОТ"""
        AdditionalPositionDialog(self.dialog, self.calculator)

    def _set_zero_headcount(self):
        if self.original_index is None: return messagebox.showwarning("Внимание", "Сначала выберите сотрудника")
        try:
            eff_month = self.calculator.months.index(self.effective_month_combo.get()) + 1
        except:
            eff_month = 1
        # ✅ ОТКРЫВАЕМ НОВЫЙ ДИАЛОГ ВМЕСТО ПРЯМОГО ОБНУЛЕНИЯ
        HeadcountEditDialog(self.dialog, self.calculator, self.original_index, eff_month - 1, gui=None)

    def _reset_changes(self):
        if self.is_adding_new:
            self.is_adding_new = False
            self.original_index = None
            self._clear_edit_form()
            self.btn_add.config(state=tk.NORMAL)
            return
        if self.original_index is None: return
        if messagebox.askyesno("Подтверждение", "Сбросить все изменения для этого сотрудника?"):
            self.calculator.reset_employee_changes(self.original_index)
            self.dialog.destroy()

    def _start_add_mode(self):
        self.is_adding_new = True
        self.original_index = None
        self.search_fio_entry.delete(0, tk.END)
        self.search_position_entry.delete(0, tk.END)
        for i in self.tree.get_children(): self.tree.delete(i)
        self.item_to_employee.clear()
        self._show_edit_form(self.calculator.create_empty_employee())
        self.btn_save.config(state=tk.NORMAL)
        self.btn_reset.config(state=tk.NORMAL)
        self.btn_zero.config(state=tk.DISABLED)
        self.btn_add.config(state=tk.DISABLED)


class PayrollGUI:
    """Графический интерфейс"""

    def __init__(self, calculator):
        self.calculator = calculator
        self.root = tk.Tk()
        self.root.title("Расчет ФОТ 2026")
        self.root.geometry("1200x700")
        try:
            icon_path = os.path.join(os.path.dirname(__file__), "icon.ico")
            if os.path.exists(icon_path):
                self.root.iconbitmap(icon_path)
        except Exception:
            pass

        self.results_text = ""
        self.last_calculation_data = None
        self._create_widgets()

    def _on_mousewheel_results(self, event):
        try:
            if hasattr(self, 'txt') and self.txt.winfo_exists():
                if event.delta:
                    self.txt.yview_scroll(int(-1 * (event.delta / 120)), "units")
                elif event.num == 4:
                    self.txt.yview_scroll(-1, "units")
                elif event.num == 5:
                    self.txt.yview_scroll(1, "units")
        except Exception:
            pass

    def _create_widgets(self):
        mf = ttk.Frame(self.root, padding="10")
        mf.pack(fill=tk.BOTH, expand=True)
        lp = ttk.Frame(mf, width=350)
        lp.pack(side=tk.LEFT, fill=tk.Y, padx=(0, 10))
        lp.pack_propagate(False)
        ttk.Button(lp, text="Загрузить Файл", command=self._load_file).pack(fill=tk.X, pady=(0, 5))
        ttk.Button(lp, text="Добавить изменение", command=self._add_change).pack(fill=tk.X, pady=(0, 10))
        hist_frame = ttk.LabelFrame(lp, text="Журнал изменений", padding=5)
        hist_frame.pack(fill=tk.BOTH, expand=True)
        self.history_list = tk.Listbox(hist_frame, height=23, font=("Arial", 9), width=40)
        self.history_list.pack(fill=tk.BOTH, expand=True)
        self.history_list.bind("<Double-1>", self._show_log_details)
        rp = ttk.Frame(mf)
        rp.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        rf = ttk.LabelFrame(rp, text="Результаты:    ", padding=10)
        rf.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        txt_frame = ttk.Frame(rf)
        txt_frame.pack(fill=tk.BOTH, expand=True)
        self.txt = tk.Text(txt_frame, wrap=tk.WORD, font=("Courier New", 9))
        self.txt.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self._txt_scrollbar = ttk.Scrollbar(txt_frame, orient=tk.VERTICAL, command=self.txt.yview)
        self._txt_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.txt.configure(yscrollcommand=self._txt_scrollbar.set)
        self.txt.bind("<MouseWheel>", self._on_mousewheel_results)
        self.txt.bind("<Button-4>", self._on_mousewheel_results)
        self.txt.bind("<Button-5>", self._on_mousewheel_results)
        bp = ttk.Frame(rp)
        bp.pack(fill=tk.X, side=tk.BOTTOM)
        ttk.Button(bp, text="Сохранить в Excel", command=self._save_to_excel).pack(side=tk.RIGHT)
        ttk.Button(bp, text="Рассчитать", command=self._calculate).pack(side=tk.RIGHT, padx=(0, 10))
        self._update_results()

    def _load_file(self):
        fp = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx *.xls")])
        if fp:
            try:
                self.calculator.load_employee_data(fp)
                messagebox.showinfo("Успех", f"Загружено сотрудников: {len(self.calculator.employees_data)}")
                self._refresh_history()
                self._update_results()
            except Exception as e:
                messagebox.showerror("Ошибка", str(e))

    def _add_change(self):
        if not self.calculator.employees_data: return messagebox.showwarning("Внимание", "Сначала загрузите файл")
        EmployeeEditDialog(self.root, self.calculator)
        self._refresh_history()

    def _refresh_history(self):
        self.history_list.delete(0, tk.END)
        for idx, log in enumerate(self.calculator.changes_log, 1):
            action = log.get('action', '')
            fio = log.get('fio', '')
            month = log.get('month', 1)
            month_name = self.calculator.months[month - 1] if 1 <= month <= 12 else str(month)
            if action == 'Изменен':
                text = f"{idx}. {action}: {fio} (с {month_name})"
            elif action == 'Выведена позиция':
                text = f"{idx}. {action}: {fio} ({month_name} → 0 чел)"
            elif action == 'Добавлена позиция':
                fot = log.get('fot_total', 0)
                text = f"{idx}. {action}: {fio} (ФОТ: {fot:,.2f} руб.)"
            elif action == 'Добавлен сотрудник':
                text = f"{idx}. {action}: {fio}"
            else:
                text = f"{idx}. {action}"
            self.history_list.insert(tk.END, text)

    def _show_log_details(self, event):
        selection = self.history_list.curselection()
        if not selection: return
        index = selection[0]
        if 0 <= index < len(self.calculator.changes_log):
            ChangeLogDetailDialog(self.root, self.calculator.changes_log[index], self.calculator, gui=self)

    def _calculate(self):
        if not self.calculator.employees_data:
            return messagebox.showwarning("Внимание", "Загрузите файл перед расчетом")
        try:
            text = "=" * 80 + "\nОБЩИЙ РАСЧЕТ ФОТ\n" + "=" * 80 + "\n\n"
            total = 0.0
            employee_results = []
            for idx in range(len(self.calculator.employees_data)):
                cur = self.calculator.get_employee_for_calculation(idx)
                emp_fot = self.calculator.calculate_total_fot_for_employee(idx)
                total += emp_fot
                marker = "*   " if idx in self.calculator.modified_employees else "    "
                comment = cur.get('comment', '')
                comment_str = f"    ({comment})" if comment else ""
                text += f"{marker}{cur['fio']:<40} {emp_fot:>15,.2f}{comment_str}\n"
                employee_results.append({'index': idx, 'fio': cur['fio'], 'fot': emp_fot})
            # ✅ Добавление ФОТ дополнительных позиций
            additional_fot = self.calculator.get_total_additional_fot()
            if additional_fot > 0:
                text += "\n" + "=" * 80 + "\nДОПОЛНИТЕЛЬНЫЕ ПОЗИЦИИ\n" + "=" * 80 + "\n\n"
                for pos in self.calculator.additional_positions:
                    pos_name = pos['position_name']
                    pos_fot = pos['fot_calculation']['total_fot']
                    total += pos_fot
                    text += f"  + {pos_name:<38} {pos_fot:>15,.2f}\n"
                    employee_results.append({'index': -1, 'fio': pos_name, 'fot': pos_fot, 'is_additional': True})
            text += "=" * 80 + f"\nИТОГО ТЕКУЩИЙ ФОТ:  {total:>15,.2f}\n"
            base_fot = self.calculator.base_fot_total
            deviation = total - base_fot
            text += f"БАЗОВЫЙ ФОТ:        {base_fot:>15,.2f}\n"
            text += f"ОТКЛОНЕНИЕ:         {deviation:>+15,.2f}\n"
            if deviation > 0:
                text += f"\n(Увеличение ФОТ на {deviation:,.2f} руб. относительно базового)\n"
            elif deviation < 0:
                text += f"\n(Уменьшение ФОТ на {abs(deviation):,.2f} руб. относительно базового)\n"
            else:
                text += f"\n(ФОТ не изменился относительно базового)\n"
            self.txt.delete(1.0, tk.END)
            self.txt.insert(tk.END, text)
            self.results_text = text
            self.last_calculation_data = {
                'total': total,
                'base_fot': base_fot,
                'deviation': deviation,
                'employees': employee_results,
                'additional_fot': additional_fot
            }
            # ✅ Всплывающее окно после расчета удалено
        except Exception as e:
            messagebox.showerror("Ошибка", str(e))
            import traceback
            traceback.print_exc()

    def _update_results(self):
        self.txt.delete(1.0, tk.END)
        self.txt.insert(tk.END, "Загрузите файл и нажмите 'Рассчитать'.")

    def _save_to_excel(self):
        if not self.last_calculation_data:
            return messagebox.showwarning("Внимание", "Сначала выполните расчет")
        save_dir = filedialog.askdirectory(title="Выберите папку для сохранения")
        if not save_dir: return
        try:
            self._save_summary_fot(save_dir)
            self._save_detailed_fot(save_dir)
            messagebox.showinfo("Успех", f"Файлы сохранены в:\n{save_dir}")
        except Exception as e:
            messagebox.showerror("Ошибка сохранения", str(e))

    def _save_summary_fot(self, save_dir):
        """Сохранение Сводный_ФОТ_2026.xlsx"""
        changed_indices = list(self.calculator.modified_employees.keys())
        if not changed_indices and not self.calculator.additional_positions:
            messagebox.showwarning("Внимание", "Нет сотрудников с изменениями. Сводный файл не будет создан.")
            return
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Сводный ФОТ (изменения)"
        headers = ['Позиция', 'ФИО', 'Показатель'] + self.calculator.months + ['Итого за год']
        ws.append(headers)
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            cell.alignment = openpyxl.styles.Alignment(horizontal='center')
        for idx in changed_indices:
            emp = self.calculator.get_employee_for_calculation(idx)
            position = emp.get('position', '')
            fio = emp.get('fio', '')
            monthly_data = [self.calculator.calculate_monthly_fot_detailed(m, employee_index=idx) for m in range(12)]
            for label, key_fn in [
                ('ФОТ ГРОСС', lambda d: d['monthly_salary'] + d['monthly_bonus']),
                ('СВ', lambda d: d['social_contributions']),
                ('Месячная премия', lambda d: d['monthly_bonus']),
                ('Резерв кварт. премий', lambda d: d['quarterly_reserve']),
                ('Резерв год. премий', lambda d: d['annual_reserve']),
                ('Компенсации', lambda d: d['compensations']),
                ('ФОТ ИТОГО', lambda d: d['fot_total']),
            ]:
                row = [position, fio, label]
                total = 0.0
                for m in range(12):
                    val = key_fn(monthly_data[m])
                    row.append(val)
                    total += val
                row.append(total)
                ws.append(row)
            ws.append([])
        if self.calculator.additional_positions:
            ws.append([])
            for pos in self.calculator.additional_positions:
                pos_name = pos['position_name']
                fot_calc = pos['fot_calculation']
                row = [pos_name, '', 'ФОТ ИТОГО']
                monthly_fot = fot_calc['monthly_fot']
                for m in range(12):
                    if m < fot_calc['months_count']:
                        row.append(monthly_fot)
                    else:
                        row.append(0)
                row.append(fot_calc['total_fot'])
                ws.append(row)
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 20
        for col in range(4, 16):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
        wb.save(os.path.join(save_dir, "Сводный_ФОТ_2026.xlsx"))

    def _save_detailed_fot(self, save_dir):
        """Сохранение Расчет_ФОТ_2026.xlsx по шаблону"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Расчет ФОТ 2026"
        headers = [
            'Дирекция по ШР', 'Юр. Лицо', 'Название МВЗ', 'Управление / департамент по ШР',
            'Отдел по ШР', 'Группа по ШР', 'Должность', 'ФИО', 'Дата (ввода, вывода)',
            'ФИО последнего занимающего должность', 'Город', '2026 начисления',
            'Оклад (1) / ЧТС (0)', 'Оклад/ЧТС по ШР, руб. (1)', 'Оклад/ЧТС по ШР, руб. (2)',
            'Доплата к окладу, руб, категория 1', 'Доплата к окладу, руб, категория 2',
            'Надбавка за проезд', 'РК, %', 'СН, %', 'Доплата за вредность, %', 'Доплата за ночные, %',
            'Компенсация за использование личной оргтехники', 'Аренда транспортного средства',
            'Аренда экипажа (авт.)', 'Лимит топливной карты (лето)', 'Лимит топливной карты (зима)',
            'Расходы на транспорт', 'Компенсация за использование личного автомобиля',
            'Лимит сотовой связи', 'Лимит интернет-связи', 'ДМС', 'Компенсация аренды жилья',
            'Компенсация проезда', 'Компенсации (лето)', 'Компенсации (зима)',
            'Месячная мотивация', 'Квартальная премия, %', 'Годовая премия, %',
            'Кол-во месяцев для расчета премии за 4 кв. LY', 'Номинальная годовая премия за LY',
            'Кол-во месяцев для расчета годовой премии LY']
        for m in self.calculator.months: headers.append(f'Числ-ть /{self.calculator.months.index(m) + 1:02d} {m} /2026')
        headers.append('Средняя численность /2026')
        headers.append('Оклад /декабря /2025')
        for m in self.calculator.months: headers.append(f'Оклад /{self.calculator.months.index(m) + 1:02d} {m} //2026')
        headers.append('Итого оклад, включая доплаты /2026')
        for m in self.calculator.months: headers.append(
            f'Премия месячная /{self.calculator.months.index(m) + 1:02d} {m} /2026')
        headers.append('Мес. Премия /2026')
        headers.extend(['Доход 01 янв', 'Доход 01 янв-02 фев', 'Доход 01 янв-03 мар',
                        'Доход 01 янв-03 мар (с кварт. Премией)', 'Доход 01 янв-04 апр', 'Доход 01 янв-05 май',
                        'Доход 01 янв-05 май (с кварт. Премией Q1 и ГБ LY)', 'Доход 01 янв-06 июн',
                        'Доход 01 янв-07 июл', 'Доход 01 янв-08 авг', 'Доход 01 янв-08 авг (с кварт. Премией Q2)',
                        'Доход 01 янв-09 сен', 'Доход 01 янв-10 окт', 'Доход 01 янв-11 ноя',
                        'Доход 01 янв-11 ноя (с кварт. Премией Q3)', 'Доход 01 янв-12 дек'])
        for m in self.calculator.months: headers.append(f'СВ /{self.calculator.months.index(m) + 1:02d} {m} /2026')
        headers.append('СВ (без резервов премий) /2026')
        for m in self.calculator.months: headers.append(
            f'Компенсации /{self.calculator.months.index(m) + 1:02d} {m} /2026')
        headers.append('Компенсации /2026')
        for m in self.calculator.months: headers.append(
            f'резервы квартальных премий /{self.calculator.months.index(m) + 1:02d} {m} /2026')
        headers.append('резервы квартальных премий /2026')
        for m in self.calculator.months: headers.append(
            f'резервы годовых премий /{self.calculator.months.index(m) + 1:02d} {m} /2026')
        headers.append('резервы годовых премий /2026')
        for m in self.calculator.months: headers.append(f'ФОТ /{self.calculator.months.index(m) + 1:02d} {m} /2026')
        headers.append('ФОТ /2026')
        headers.extend(['ФОТ', 'СВ', 'Комментарий', 'ФОТ2', 'СВ в бюджете', 'Регресс',
                        'Город базирования', 'Характер работы', 'График работы', 'Уровень', 'Город'])
        ws.append(headers)
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True, size=9)
            cell.fill = openpyxl.styles.PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', wrap_text=True)
        for idx in range(len(self.calculator.employees_data)):
            emp = self.calculator.get_employee_for_calculation(idx)
            monthly_data = [self.calculator.calculate_monthly_fot_detailed(m, employee_index=idx) for m in range(12)]
            row = [emp.get('directorate', ''), emp.get('legal_entity', ''), '', '', '', '',
                   emp.get('position', ''), emp.get('fio', ''), '', '', emp.get('city', ''), 2026, 1,
                   emp.get('base_salary', 0.0), 0, 0, 0, 0,
                   emp.get('rk_percent', 0.0), emp.get('sn_percent', 0.0), 0, 0,
                   0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0,
                   emp.get('compensation_summer', 0.0), emp.get('compensation_winter', 0.0),
                   emp.get('monthly_bonus_percent', 0.0), emp.get('quarterly_bonus_percent', 0.0),
                   emp.get('annual_bonus_percent', 0.0), 0, 0, 0]
            headcounts = emp.get('headcounts', [1.0] * 12)
            row.extend(headcounts)
            row.append(sum(headcounts) / 12)
            row.append(0)
            base_salary_monthly = emp.get('base_salary_monthly', [emp.get('base_salary', 0.0)] * 12)
            row.extend(base_salary_monthly)
            row.append(sum(base_salary_monthly))
            row.extend([monthly_data[m]['monthly_bonus'] for m in range(12)])
            row.append(sum(monthly_data[m]['monthly_bonus'] for m in range(12)))
            row.extend([0] * 16)
            row.extend([monthly_data[m]['social_contributions'] for m in range(12)])
            row.append(sum(monthly_data[m]['social_contributions'] for m in range(12)))
            row.extend([monthly_data[m]['compensations'] for m in range(12)])
            row.append(sum(monthly_data[m]['compensations'] for m in range(12)))
            row.extend([monthly_data[m]['quarterly_reserve'] for m in range(12)])
            row.append(sum(monthly_data[m]['quarterly_reserve'] for m in range(12)))
            row.extend([monthly_data[m]['annual_reserve'] for m in range(12)])
            row.append(sum(monthly_data[m]['annual_reserve'] for m in range(12)))
            fot_monthly = [monthly_data[m]['fot_total'] for m in range(12)]
            row.extend(fot_monthly)
            row.append(sum(fot_monthly))
            row.append(sum(fot_monthly))
            row.append(sum(monthly_data[m]['social_contributions'] for m in range(12)))
            row.append(emp.get('comment', ''))
            row.extend(['', '', '', '', '', '', ''])
            ws.append(row)
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['G'].width = 40
        ws.column_dimensions['H'].width = 35
        wb.save(os.path.join(save_dir, "Расчет_ФОТ_2026.xlsx"))

    def run(self):
        self.root.mainloop()


def run_gui():
    calculator = PayrollCalculator()
    app = PayrollGUI(calculator)
    app.run()


if __name__ == "__main__":
    run_gui()